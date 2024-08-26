import pandas as pd
import matplotlib.pyplot as plt
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
import os
from reportlab.lib.units import inch, cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import numpy as np
import ezdxf

def load_data(csv_file):
    return pd.read_csv(csv_file)

def check_file_type(file_path):
    try:
        doc = ezdxf.readfile(file_path)
        for entity in doc.entities:
            if entity.dxftype() == "TEXT" or entity.dxftype() == "MTEXT":
                text = entity.dxf.text.lower()
                if ('danfoss' in text or '16016' in text) and 'baden' not in text:
                    return 'Danfoss file detected'
                elif ('danfoss' in text or '16016' in text) and 'baden' in text:
                    return 'Aeroquip file detected'
                elif 'aeroquip' in text:
                    return 'Aeroquip file detected'
                elif 'eaton' in text:
                    return 'Eaton file detected'
                elif 'obsolete' in text:
                    return 'Obsolete file detected'
    except Exception as e:
        print(f"Error reading file {file_path}: {str(e)}")
    return ''

def determine_result(row):
    # Check if filename starts with '1EW-' or '1BW-'
    if row['filename'].startswith(('1EW-', '1BW-')):
        return "Unmodified / Review"
    
    # First, check for Danfoss file
    if "Danfoss file detected" in str(row['issue']):
        return "Danfoss File"
    
    # Then check for modifications
    if row['logo'] > 0 or row['division'] > 0 or row['ip_change'] > 0 or row['text'] > 1:
        return "Modified file (elements changed)"
    
    # Other conditions
    if "Old IP w/ Word Chg" in row['possible_errors'] or "2 lines of IP2 in Bottom" in row['possible_errors']:
        return "Unmodified / Send to Review"
    elif row['logo'] == 0 and row['division'] == 0 and row['ip_change'] == 0 and (row['text'] == 0 or row['text'] == 1):
        return "Unmodified / Review"
    elif "All extra checks passed" in str(row['extra_checks']) and "File size changed by more than 15%" in str(row['issue']):
        return "Modified"
    elif "Send to Review" in str(row['issue']) or "Aeroquip file detected" in str(row['issue']):
        return "Unmodified / Review"
    else:
        return "Unmodified / Review"
    
    
def check_dxf_conditions(file_path):
    try:
        doc = ezdxf.readfile(file_path)
        modelspace = doc.modelspace()
        
        # Define the areas to check (adjust these values as needed)
        top_left_area = (0, doc.header['$EXTMAX'].y * 0.8, doc.header['$EXTMAX'].x * 0.2, doc.header['$EXTMAX'].y)
        bottom_left_area = (0, 0, doc.header['$EXTMAX'].x * 0.2, doc.header['$EXTMAX'].y * 0.2)
        
        top_left_text = []
        bottom_left_text = []
        
        for entity in modelspace:
            if entity.dxftype() in ('TEXT', 'MTEXT'):
                x, y = entity.dxf.insert
                
                if top_left_area[0] <= x <= top_left_area[2] and top_left_area[1] <= y <= top_left_area[3]:
                    top_left_text.append(entity.dxf.text)
                
                if bottom_left_area[0] <= x <= bottom_left_area[2] and bottom_left_area[1] <= y <= bottom_left_area[3]:
                    bottom_left_text.append(entity.dxf.text)
        
        result = []
        
        # Check top-left text
        expected_top_left_text = "THE REPRODUCTION, DISTRIBUTION, AND UTILIZATION OF THIS DOCUMENT AS WELL AS THE COMMUNICATION OF ITS CONTENTS TO OTHERS WITHOUT EXPLICIT AUTHORIZATION IS PROHIBITED. OFFENDERS WILL BE HELD LIABLE FOR THE PAYMENT OF DAMAGES. ALL RIGHTS RESERVED INTHE EVENT OF THE GRANT OF A PATENT, UTILITY MODEL OR DESIGN. (PER ISO 16016)"
        if top_left_text and top_left_text[0] != expected_top_left_text:
            result.append("Old IP w/ Word Chg")
        
        # Check bottom-left text
        if len(bottom_left_text) == 2:
            result.append("2 lines of IP2 in Bottom")
        
        return ', '.join(result) if result else "No doubts"
    
    except Exception as e:
        print(f"Error checking DXF file {file_path}: {str(e)}")
        return "Error checking file"


def generate_insights(df, mod_dir):
    total_files = len(df)
    logos_detected = df['logo'].sum()
    divisions_detected = df['division'].sum()
    texts_detected = df['text'].sum()
    ip_changes_detected = df['ip_change'].sum()
    
    # Create a copy of the dataframe for modifications
    files_with_changes = df.copy()
    
    # Add an 'issue' column
    files_with_changes['issue'] = ''
    
    # Check file type for each file in the modified directory
    files_with_changes['file_type'] = files_with_changes['filename'].apply(lambda x: check_file_type(os.path.join(mod_dir, x)))
    
    # Identify different types of files
    modified_not_found = files_with_changes['comments'].str.contains('Modified file not found', na=False)
    large_size_changes = files_with_changes['comments'].str.contains('File size changed by more than 15%', na=False)
    
    # Calculate total changes
    files_with_changes['total_changes'] = files_with_changes['logo'] + files_with_changes['division'] + files_with_changes['text'] + files_with_changes['ip_change']
    #Results
    
    files_with_changes['possible_errors'] = files_with_changes['filename'].apply(
        lambda x: check_dxf_conditions(os.path.join(mod_dir, x)) if x.startswith(('1JM','1JH','247', '144')) else "No doubts"
    )

    # Update the 'issue' column
    files_with_changes.loc[files_with_changes['file_type'] != '', 'issue'] = files_with_changes['file_type']
    files_with_changes.loc[modified_not_found, 'issue'] = "Modified file not found"
    files_with_changes.loc[large_size_changes, 'issue'] = "File size changed by more than 15% -> Modified"
    
    # Identify files with no changes and no significant block changes
    no_changes = (files_with_changes['total_changes'] == 0) & (files_with_changes['block_changes'].str.contains('No significant block changes detected', na=False))
    files_with_changes.loc[no_changes & (files_with_changes['issue'] == ''), 'issue'] = "Send to Review"
    
    # Identify files with changes but no significant size change
    small_changes = (files_with_changes['total_changes'] >= 2) & (~large_size_changes)
    files_with_changes.loc[small_changes & (files_with_changes['issue'] == ''), 'issue'] = "Modified file (elements changed)"
    
    # For remaining files with no issue identified, mark as "Send to Review"
    files_with_changes.loc[files_with_changes['issue'] == '', 'issue'] = "Send to Review"
    
    def determine_result(row):
    # Check if filename starts with '1EW-' or '1BW-'
        if row['filename'].startswith(('1EW-', '1BW-')):
            return "Unmodified / Review"
        
        # First, check for Danfoss file
        if "Danfoss file detected" in str(row['issue']):
            return "Danfoss File"
        
        # Then check for modifications
        if row['logo'] > 0 or row['division'] > 0 or row['ip_change'] > 0 or row['text'] > 1:
            return "Modified file (elements changed)"
        
        # Other conditions
        if "Old IP w/ Word Chg" in row['possible_errors'] or "2 lines of IP2 in Bottom" in row['possible_errors']:
            return "Unmodified / Send to Review"
        elif row['logo'] == 0 and row['division'] == 0 and row['ip_change'] == 0 and (row['text'] == 0 or row['text'] == 1):
            return "Unmodified / Review"
        elif "All extra checks passed" in str(row['extra_checks']) and "File size changed by more than 15%" in str(row['issue']):
            return "Modified"
        elif "Send to Review" in str(row['issue']) or "Aeroquip file detected" in str(row['issue']):
            return "Unmodified / Review"
        else:
            return "Unmodified / Review"
    files_with_changes['result'] = files_with_changes.apply(determine_result, axis=1)
    
    # Other insights
    files_with_multiple_logos = files_with_changes[files_with_changes['logo'] > 1]
    files_with_multiple_ip_notes = files_with_changes[files_with_changes['ip_change'] > 1]
    files_with_changes['block_changes'] = files_with_changes['block_changes'].fillna('').astype(str)
    files_with_extra_checks = files_with_changes[files_with_changes['extra_checks'] != 'All extra checks passed']
    files_sent_to_review = len(files_with_changes[files_with_changes['issue'] == "Send to Review"])

    return {
        'total_files': total_files,
        'logos_detected': logos_detected,
        'divisions_detected': divisions_detected,
        'texts_detected': texts_detected,
        'ip_changes_detected': ip_changes_detected,
        'files_with_changes': files_with_changes,
        'files_with_multiple_logos': files_with_multiple_logos,
        'files_with_multiple_ip_notes': files_with_multiple_ip_notes,
        'files_with_extra_checks': files_with_extra_checks,
        'files_sent_to_review': files_sent_to_review
    }

def compare_with_ideal(df, ideal_dir):
    ideal_files = set(os.listdir(ideal_dir))
    processed_files = set(df['filename'])
    
    similar_to_ideal = len(ideal_files.intersection(processed_files))
    different_from_ideal = len(processed_files - ideal_files)
    
    return similar_to_ideal, different_from_ideal

def calculate_accuracy(insights):
    # total_files = insights['total_files']
    # danfoss_files = len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Danfoss file detected'])
    # non_modified_files = len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Modified file not found'])
    # processable_files = total_files - danfoss_files - non_modified_files
    
    total_files = insights['total_files']
    danfoss_files = len(insights['files_with_changes'][insights['files_with_changes']['result'] == 'Danfoss File'])
    non_modified_files = len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Modified file not found'])
    processable_files = total_files - danfoss_files - non_modified_files


    logo_detections = insights['logos_detected']
    division_detections = insights['divisions_detected']
    text_detections = insights['texts_detected']
    ip_change_detections = insights['ip_changes_detected']

    total_detections = logo_detections + division_detections + text_detections + ip_change_detections
    expected_detections = processable_files * 2  

    base_accuracy = (total_detections / expected_detections) * 100 if expected_detections > 0 else 0

    extra_check_failures = len(insights['files_with_extra_checks'])
    extra_check_penalty = (extra_check_failures / processable_files) * 100 if processable_files > 0 else 0

    accuracy = max(0, base_accuracy - extra_check_penalty)

    return accuracy

def create_pdf_report(insights, similar_to_ideal, different_from_ideal, accuracy, output_file):
    doc = SimpleDocTemplate(output_file, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch)

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Quality Check Analysis Report", styles['Title']))
    elements.append(Spacer(1, 12))
    
    total_files = insights['total_files']
    files_sent_to_review = insights['files_sent_to_review']
    new_accuracy = ((total_files - files_sent_to_review) / total_files) * 100 if total_files > 0 else 0

    # General Insights
    data = [
        ["Total Files Processed", insights['total_files']],
        ["Logos Detected", insights['logos_detected']],
        ["Divisions Detected", insights['divisions_detected']],
        ["Texts Detected", insights['texts_detected']],
        ["IP Changes Detected", insights['ip_changes_detected']],
        ["Danfoss Files Detected", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Danfoss file detected'])],
        ["Modified Files Not Found", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Modified file not found'])],
        ["Files with Changes", len(insights['files_with_changes'][insights['files_with_changes']['logo'] + insights['files_with_changes']['division'] + insights['files_with_changes']['text'] + insights['files_with_changes']['ip_change'] > 0])],
        ["Files with Large Size Changes", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'File size changed by more than 15% -> Modified'])],
        ["Files with Few Changes", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'No changes detected'])],
        ["Files Similar to Ideal", similar_to_ideal],
        ["Files Different from Ideal", different_from_ideal],
        #["Accuracy (Files not sent to review)", f"{new_accuracy:.2f}%"],  

    ]

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    elements.append(PageBreak())

    # Files to Review
    elements.append(Paragraph("Files Review Detail", styles['Heading1']))
    elements.append(Spacer(1, 12))

    review_data = [["Filename", "Issue", "Changes", "Block Changes", "Extra Checks in DWG", "Result"]]

    wrapped_style = ParagraphStyle('WrappedStyle', parent=styles['Normal'], fontSize=6, leading=8)

    for _, row in insights['files_with_changes'].iterrows():
        #changes = row['logo'] + row['division'] + row['text'] + row['ip_change']
        changes = row['total_changes']
        change_text = f"{changes} elements changed" if changes > 0 else row['issue']
        
        cell_style = wrapped_style
        if changes > 4:
            cell_style = ParagraphStyle('HighlightedStyle', parent=wrapped_style, backColor=colors.yellow)
        result = determine_result(row)  # Now this will work

        review_data.append([
            Paragraph(row['filename'], wrapped_style),
            Paragraph(row['issue'], wrapped_style),
            Paragraph(change_text, cell_style),
            Paragraph(str(row['block_changes']), wrapped_style),
            Paragraph(str(row['extra_checks']), wrapped_style),
            Paragraph(str(row['result']), wrapped_style)
        ])

    col_widths = [2.5*cm, 2*cm, 2*cm, 5*cm, 3.5*cm, 2*cm]  # Adjust column widths

    review_table = Table(review_data, colWidths=col_widths)
    review_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(review_table)

    doc.build(elements)

def save_tables_to_excel(insights, similar_to_ideal, different_from_ideal, accuracy, csv_file):
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    row_fill = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid")  # Light pink

    # Create "General Insights" sheet
    ws1 = wb.active
    ws1.title = "General Insights"
    
    total_files = insights['total_files']
    files_sent_to_review = insights['files_sent_to_review']
    new_accuracy = ((total_files - files_sent_to_review) / total_files) * 100 if total_files > 0 else 0
    
    data = [
        ["Total Files Processed", insights['total_files']],
        ["Logos Detected", insights['logos_detected']],
        ["Divisions Detected", insights['divisions_detected']],
        ["Texts Detected", insights['texts_detected']],
        ["IP Changes Detected", insights['ip_changes_detected']],
        ["Danfoss Files Detected", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Danfoss file detected'])],
        ["Modified Files Not Found", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'Modified file not found'])],
        ["Files with Changes", len(insights['files_with_changes'][insights['files_with_changes']['logo'] + insights['files_with_changes']['division'] + insights['files_with_changes']['text'] + insights['files_with_changes']['ip_change'] > 0])],
        ["Files with Large Size Changes", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'File size changed by more than 15% -> Modified'])],
        ["Files with Few Changes", len(insights['files_with_changes'][insights['files_with_changes']['issue'] == 'No changes detected'])],
        ["Files Similar to Ideal", similar_to_ideal],
        ["Files Different from Ideal", different_from_ideal],
        #["Accuracy (Files not sent to review)", f"{new_accuracy:.2f}%"],  # Add this line

    ]
    
    for row in data:
        ws1.append(row)
    
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.fill = row_fill
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # Create "Files to Review" sheet
    ws2 = wb.create_sheet("Files Review Detail")
    
    review_data = [["Result", "Filename", "Logo", "Division", "Text", "IP Change", "Issue", "Changes", "Block Changes", "Extra Checks in DWG", "Possible Errors"]]
    
    for _, row in insights['files_with_changes'].iterrows():
        #changes = row['logo'] + row['division'] + row['text'] + row['ip_change']
        changes = row['total_changes']
        change_text = f"{changes} elements changed" if changes > 0 else row['issue']
        result = determine_result(row)  # Use the same function to determine the result
        review_data.append([
            row['filename'],
            row['result'],
            row['logo'],
            row['division'],
            row['text'],
            row['ip_change'],
            row['issue'],
            change_text,
            str(row['block_changes']),
            str(row['extra_checks']),
            row['possible_errors']
            
        ])
    
    for row in review_data:
        ws2.append(row)
        
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for cell in row:
            cell.fill = row_fill
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # Create "Original CSV Data" sheet
    ws3 = wb.create_sheet("Original CSV Data")
    
    # Read the original CSV file
    df = pd.read_csv(csv_file)
    
    # Write the column headers
    ws3.append(df.columns.tolist())
    
    # Write the data rows
    for _, row in df.iterrows():
        ws3.append(row.tolist())
    
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        for cell in row:
            cell.fill = row_fill
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
                
    for ws in [ws1, ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    excel_file = csv_file.replace('.csv', '_with_insights.xlsx')
    wb.save(excel_file)
    print(f"Excel file with insights saved: {excel_file}")

def main():
    csv_file = 'ComparisonOutput\comprehensive_quality_check_run_test_batch_beyond97k.csv'
    ideal_dir = 'IdealDXF'
    output_file = 'quality_check_report_batch_97k_1.pdf'
    src_dir = 'Beyond_97k_DXF_In' 
    mod_dir = 'Beyond_97k_DXF_Out'

    df = load_data(csv_file)
    insights = generate_insights(df, mod_dir)
    similar_to_ideal, different_from_ideal = compare_with_ideal(df, ideal_dir)
    accuracy = calculate_accuracy(insights)

    create_pdf_report(insights, similar_to_ideal, different_from_ideal, accuracy, output_file)
    print(f"PDF report generated: {output_file}")
    
    save_tables_to_excel(insights, similar_to_ideal, different_from_ideal, accuracy, csv_file)

if __name__ == '__main__':
    main()